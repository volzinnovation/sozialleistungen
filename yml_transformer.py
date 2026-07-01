#!/usr/bin/env python3
"""Generate CSV, Excel, JSON, and a static explorer from sozialleistungen.yml."""

from __future__ import annotations

import argparse
import csv
import html
import json
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
import yaml


ROOT = Path(__file__).resolve().parent
DEFAULT_YAML = ROOT / "sozialleistungen.yml"
DEFAULT_CSV = ROOT / "sozialleistungen.csv"
DEFAULT_XLSX = ROOT / "sozialleistungen.xlsx"
DEFAULT_JSON = ROOT / "sozialleistungen.json"
DEFAULT_HTML = ROOT / "index.html"

FIELDNAMES = [
    "leistung",
    "rechtsnorm",
    "zielgruppen",
    "themenfelder",
    "category",
    "sub_category",
    "zielgruppe1",
    "zielgruppe2",
    "themenfeld1",
    "themenfeld2",
]


def as_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value).strip()
    return [text] if text else []


def split_topics(values: list[str]) -> list[str]:
    topics: list[str] = []
    for value in values:
        topics.extend(part.strip() for part in value.split("&") if part.strip())
    return topics


def load_rows(source: Path = DEFAULT_YAML) -> list[dict[str, str]]:
    payload = yaml.safe_load(source.read_text(encoding="utf-8"))
    rows: list[dict[str, str]] = []
    for category, sub_categories in payload.items():
        for sub_category, items in sub_categories.items():
            for item in items:
                target_groups = as_list(item.get("zielgruppen"))
                topic_groups = as_list(item.get("themenfelder"))
                topics = split_topics(topic_groups)
                rows.append(
                    {
                        "leistung": str(item.get("leistung", "")).strip(),
                        "rechtsnorm": str(item.get("rechtsnorm", "")).strip(),
                        "zielgruppen": repr(target_groups),
                        "themenfelder": repr(topic_groups),
                        "category": str(category),
                        "sub_category": str(sub_category),
                        "zielgruppe1": target_groups[0] if len(target_groups) > 0 else "",
                        "zielgruppe2": target_groups[1] if len(target_groups) > 1 else "",
                        "themenfeld1": topics[0] if len(topics) > 0 else "",
                        "themenfeld2": topics[1] if len(topics) > 1 else "",
                    }
                )
    return rows


def write_csv(rows: list[dict[str, str]], target: Path = DEFAULT_CSV) -> None:
    with target.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def current_xlsx_values(path: Path) -> list[list[str]]:
    if not path.exists():
        return []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = [
        ["" if value is None else str(value) for value in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    workbook.close()
    return values


def write_xlsx_if_changed(rows: list[dict[str, str]], target: Path = DEFAULT_XLSX) -> None:
    desired = [FIELDNAMES] + [[row[field] for field in FIELDNAMES] for row in rows]
    if current_xlsx_values(target) == desired:
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sozialleistungen"
    sheet.append(FIELDNAMES)
    for row in rows:
        sheet.append([row[field] for field in FIELDNAMES])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(target)


def payload_for(rows: list[dict[str, str]]) -> dict[str, Any]:
    category_counts = Counter(row["category"] for row in rows)
    target_counts = Counter(row["zielgruppe1"] for row in rows if row["zielgruppe1"])
    topic_counts = Counter(row["themenfeld1"] for row in rows if row["themenfeld1"])
    return {
        "generated_by": "yml_transformer.py",
        "row_count": len(rows),
        "category_count": len(category_counts),
        "target_group_count": len(target_counts),
        "topic_count": len(topic_counts),
        "categories": dict(sorted(category_counts.items())),
        "target_groups": dict(sorted(target_counts.items())),
        "topics": dict(sorted(topic_counts.items())),
        "rows": rows,
    }


def write_json(rows: list[dict[str, str]], target: Path = DEFAULT_JSON) -> None:
    target.write_text(
        json.dumps(payload_for(rows), ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def options_for(values: list[str]) -> str:
    return "".join(
        f'<option value="{html.escape(value, quote=True)}">{html.escape(value)}</option>'
        for value in values
    )


def render_table_rows(rows: list[dict[str, str]]) -> str:
    rendered = []
    for row in rows:
        search = html.escape(" ".join(row.values()).lower(), quote=True)
        rendered.append(
            f"""
            <tr data-search="{search}" data-category="{html.escape(row['category'], quote=True)}"
                data-target="{html.escape(row['zielgruppe1'], quote=True)}"
                data-topic="{html.escape(row['themenfeld1'], quote=True)}">
              <td>{html.escape(row['category'])}</td>
              <td>{html.escape(row['sub_category'])}</td>
              <td>{html.escape(row['leistung'])}</td>
              <td>{html.escape(row['rechtsnorm'])}</td>
              <td>{html.escape(row['zielgruppe1'])}</td>
              <td>{html.escape(row['themenfeld1'])}</td>
            </tr>
            """.strip()
        )
    return "\n".join(rendered)


def write_html(rows: list[dict[str, str]], target: Path = DEFAULT_HTML) -> None:
    data = payload_for(rows)
    categories = sorted(data["categories"])
    targets = sorted(data["target_groups"])
    topics = sorted(data["topics"])
    target.write_text(
        f"""<!doctype html>
<html lang="de">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1" />
    <title>Sozialleistungen Explorer</title>
    <style>
      :root {{
        --bg: #f4f1ea;
        --ink: #17202a;
        --muted: #66717d;
        --line: #d8d0c3;
        --paper: #fffdf8;
        --accent: #115e59;
      }}
      * {{ box-sizing: border-box; }}
      body {{
        background: var(--bg);
        color: var(--ink);
        font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
        line-height: 1.45;
        margin: 0;
      }}
      main {{ margin: 0 auto; max-width: 1280px; padding: 2.4rem 1rem; }}
      h1 {{ font-size: clamp(2.1rem, 6vw, 4.8rem); letter-spacing: 0; line-height: 1; margin: 0; }}
      p {{ margin: 0; }}
      .lead {{ color: var(--muted); margin-top: 0.8rem; max-width: 760px; }}
      .metrics {{ display: flex; flex-wrap: wrap; gap: 0.7rem; margin: 1.4rem 0; }}
      .metrics span {{ background: var(--paper); border: 1px solid var(--line); border-radius: 999px; padding: 0.45rem 0.8rem; }}
      .filters {{ display: grid; gap: 0.7rem; grid-template-columns: minmax(220px, 1fr) repeat(3, minmax(140px, 0.45fr)); margin-bottom: 1rem; }}
      input, select, button {{ background: var(--paper); border: 1px solid var(--line); border-radius: 8px; color: var(--ink); font: inherit; min-height: 2.75rem; padding: 0.55rem 0.7rem; width: 100%; }}
      button {{ cursor: pointer; font-weight: 800; width: auto; }}
      button:hover {{ border-color: var(--accent); color: var(--accent); }}
      .table-wrap {{ background: var(--paper); border: 1px solid var(--line); border-radius: 8px; overflow: auto; }}
      table {{ border-collapse: collapse; min-width: 980px; width: 100%; }}
      th, td {{ border-bottom: 1px solid var(--line); padding: 0.65rem 0.75rem; text-align: left; vertical-align: top; }}
      th {{ background: #ebe4d8; font-size: 0.82rem; position: sticky; top: 0; z-index: 1; }}
      td {{ font-size: 0.9rem; }}
      tr[hidden] {{ display: none; }}
      .status {{ color: var(--muted); font-weight: 700; margin: 0; }}
      .filter-tools {{ align-items: center; display: flex; flex-wrap: wrap; gap: 0.7rem; justify-content: space-between; margin-bottom: 0.7rem; }}
      .copy-status {{ color: var(--muted); font-size: 0.9rem; min-height: 1.3rem; }}
      .downloads {{ display: flex; flex-wrap: wrap; gap: 0.6rem; margin-bottom: 1rem; }}
      .downloads a {{ background: var(--accent); border-radius: 6px; color: white; font-weight: 700; padding: 0.5rem 0.75rem; text-decoration: none; }}
      @media (max-width: 820px) {{
        main {{ padding: 1.6rem 0.8rem; }}
        .filters {{ grid-template-columns: 1fr; }}
      }}
    </style>
  </head>
  <body>
    <main>
      <header>
        <h1>Sozialleistungen Explorer</h1>
        <p class="lead">Durchsuchbarer statischer Explorer der YAML-Inventur. Filter greifen rein lokal im Browser; die Daten bleiben als CSV, Excel und JSON im Repository.</p>
      </header>
      <section class="metrics" aria-label="Datenumfang">
        <span><strong>{data['row_count']}</strong> Leistungen</span>
        <span><strong>{data['category_count']}</strong> Gesetzbuecher</span>
        <span><strong>{data['target_group_count']}</strong> Zielgruppen</span>
        <span><strong>{data['topic_count']}</strong> Themenfelder</span>
      </section>
      <nav class="downloads" aria-label="Downloads">
        <a href="sozialleistungen.csv">CSV</a>
        <a href="sozialleistungen.xlsx">Excel</a>
        <a href="sozialleistungen.json">JSON</a>
        <a href="sozialleistungen.yml">YAML</a>
      </nav>
      <section class="filters" aria-label="Filter">
        <input id="search" type="search" placeholder="Suche nach Leistung, Norm, Kategorie" />
        <select id="category"><option value="">Alle Gesetzbuecher</option>{options_for(categories)}</select>
        <select id="target"><option value="">Alle Zielgruppen</option>{options_for(targets)}</select>
        <select id="topic"><option value="">Alle Themenfelder</option>{options_for(topics)}</select>
      </section>
      <div class="filter-tools">
        <p class="status"><span id="visible-count">{len(rows)}</span> / {len(rows)} sichtbar</p>
        <button id="copy-filter-link" type="button">Filter-Link kopieren</button>
        <span id="copy-status" class="copy-status" aria-live="polite"></span>
      </div>
      <section class="table-wrap" aria-label="Sozialleistungen Tabelle">
        <table>
          <thead>
            <tr>
              <th>Gesetzbuch</th>
              <th>Kategorie</th>
              <th>Leistung</th>
              <th>Rechtsnorm</th>
              <th>Zielgruppe</th>
              <th>Themenfeld</th>
            </tr>
          </thead>
          <tbody>
{render_table_rows(rows)}
          </tbody>
        </table>
      </section>
    </main>
    <script>
      const rows = Array.from(document.querySelectorAll('tbody tr'));
      const search = document.getElementById('search');
      const category = document.getElementById('category');
      const target = document.getElementById('target');
      const topic = document.getElementById('topic');
      const visibleCount = document.getElementById('visible-count');
      const copyButton = document.getElementById('copy-filter-link');
      const copyStatus = document.getElementById('copy-status');
      const filterNodes = [search, category, target, topic];
      const params = new URLSearchParams(window.location.search);

      function setSelectValue(select, value) {{
        if (!value) return;
        const hasOption = Array.from(select.options).some((option) => option.value === value);
        if (hasOption) select.value = value;
      }}

      function readUrlFilters() {{
        search.value = params.get('search') || '';
        setSelectValue(category, params.get('category') || '');
        setSelectValue(target, params.get('target') || '');
        setSelectValue(topic, params.get('topic') || '');
      }}

      function currentFilterParams() {{
        const next = new URLSearchParams();
        if (search.value.trim()) next.set('search', search.value.trim());
        if (category.value) next.set('category', category.value);
        if (target.value) next.set('target', target.value);
        if (topic.value) next.set('topic', topic.value);
        return next;
      }}

      function updateUrl() {{
        const next = currentFilterParams();
        const query = next.toString();
        const nextUrl = query
          ? `${{window.location.pathname}}?${{query}}`
          : window.location.pathname;
        window.history.replaceState(null, '', nextUrl);
      }}

      function applyFilters(updateAddress = true) {{
        const query = search.value.trim().toLowerCase();
        let visible = 0;
        rows.forEach((row) => {{
          const matches = (!query || row.dataset.search.includes(query))
            && (!category.value || row.dataset.category === category.value)
            && (!target.value || row.dataset.target === target.value)
            && (!topic.value || row.dataset.topic === topic.value);
          row.hidden = !matches;
          if (matches) visible += 1;
        }});
        visibleCount.textContent = visible;
        if (updateAddress) updateUrl();
      }}

      async function writeClipboard(text) {{
        if (navigator.clipboard?.writeText) {{
          try {{
            await navigator.clipboard.writeText(text);
            return;
          }} catch (_error) {{
            // Fall back to document selection below.
          }}
        }}
        const textarea = document.createElement('textarea');
        textarea.value = text;
        textarea.setAttribute('readonly', '');
        textarea.style.position = 'fixed';
        textarea.style.top = '-999px';
        document.body.append(textarea);
        textarea.select();
        try {{
          if (!document.execCommand('copy')) throw new Error('Copy rejected');
        }} finally {{
          textarea.remove();
        }}
      }}

      async function copyFilterLink() {{
        updateUrl();
        try {{
          await writeClipboard(window.location.href);
          copyStatus.textContent = 'Filter-Link kopiert.';
          window.setTimeout(() => {{
            copyStatus.textContent = '';
          }}, 1600);
        }} catch (error) {{
          console.error(error);
          copyStatus.textContent = 'Kopieren ist in diesem Browser nicht verfügbar.';
        }}
      }}

      readUrlFilters();
      applyFilters(false);
      filterNodes.forEach((node) => node.addEventListener('input', () => applyFilters(true)));
      copyButton.addEventListener('click', copyFilterLink);
    </script>
  </body>
</html>
""",
        encoding="utf-8",
    )


def generate(source: Path = DEFAULT_YAML) -> list[dict[str, str]]:
    rows = load_rows(source)
    write_csv(rows)
    write_xlsx_if_changed(rows)
    write_json(rows)
    write_html(rows)
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_YAML)
    args = parser.parse_args()
    rows = generate(args.source)
    print(f"Generated {len(rows)} Sozialleistungen.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
